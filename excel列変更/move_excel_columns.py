import sys
from openpyxl import load_workbook


def parse_range(s: str) -> tuple[int, int]:
    start, end = s.split(":")
    return int(start), int(end)


def move_columns_all_sheets(xlsx_file: str, src_range: str, dst_range: str):
    """
    指定した列範囲を、全シートに対して移動（値のみ）する。
    例:
        python move_excel_columns.py sample.xlsx 2:3 6:7
    """

    src_start, src_end = parse_range(src_range)
    dst_start, dst_end = parse_range(dst_range)

    width = src_end - src_start + 1
    if width != (dst_end - dst_start + 1):
        raise ValueError("src と dst の列数は同じである必要があります")

    wb = load_workbook(xlsx_file)

    for ws in wb.worksheets:
        print(f"processing sheet: {ws.title}")

        # ===== 1. 元列の値を退避 =====
        cols_data = []
        for col in range(src_start, src_end + 1):
            col_cells = next(ws.iter_cols(min_col=col, max_col=col))
            col_data = [cell.value for cell in col_cells]
            cols_data.append(col_data)

        # ===== 2. 削除順序制御 =====
        # dst が src より右にある場合は、先に削除してから貼り付け
        if dst_start > src_start:
            for col in range(src_end, src_start - 1, -1):
                ws.delete_cols(col)

            # 削除後、dst 位置が左にずれる
            shift = width
            actual_dst_start = dst_start - shift
        else:
            actual_dst_start = dst_start

        # ===== 3. 貼り付け =====
        for i, col_data in enumerate(cols_data):
            for row_idx, value in enumerate(col_data, start=1):
                ws.cell(
                    row=row_idx,
                    column=actual_dst_start + i,
                    value=value,
                )

        # ===== 4. dst < src の場合は後で削除 =====
        if dst_start < src_start:
            for col in range(src_end, src_start - 1, -1):
                ws.delete_cols(col)

    wb.save(xlsx_file)
    print("done")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("usage: python move_excel_columns.py file.xlsx 2:3 6:7")
        sys.exit(1)

    move_columns_all_sheets(sys.argv[1], sys.argv[2], sys.argv[3])